from pathlib import Path

import numpy as np
import openpyxl


def load_named_range(workbook, range_name):
    rng = workbook.defined_names[range_name]
    sht, coordinates = tuple(rng.destinations)[0]
    ws = workbook[sht]
    return np.array([[cell.value for cell in row] for row in ws[coordinates]])


if __name__ == '__main__':
    wb = openpyxl.load_workbook(Path.cwd() / 'excel/workbook with named range.xlsx')
    print('Loaded {} sheets: {}'.format(len(wb.sheetnames),
                                        ', '.join(wb.sheetnames)))
    my_named_range = load_named_range(wb, 'my_named_range')
    print(my_named_range)
